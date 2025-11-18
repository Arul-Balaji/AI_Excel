import openpyxl
import pandas as pd
from datetime import datetime
import os
import shutil

class SalesForecastModel:
    """
    A class to interact with the Microsoft Sales Forecast Tracker spreadsheet using openpyxl.
    """

    def __init__(self, file_path="microsoft_Sales forecast tracker small business.xlsx"):
        """
        Initialize the SalesForecastModel with the path to the Excel file.
        """
        self.name = "sales_forecast"
        self.file_path = file_path
        if not os.path.exists(file_path):
            raise FileNotFoundError(f"Excel file not found: {file_path}")

    def _get_last_row(self, sheet, column_letter):
        """Helper to find the last row with data in a given column, starting from the bottom."""
        for cell in sheet[column_letter][::-1]:
            # Data starts below row 6
            if cell.row > 6 and cell.value is not None:
                return cell.row
        return 6 # If no data, return the header row

    def add_forecast_input_row(self, data_dict):
        """
        Add a new row to the 'Forecast input' tab using openpyxl.
        """
        wb = openpyxl.load_workbook(self.file_path)
        input_sheet = wb['Forecast input']

        # Headers are in B6:J6 (columns 2 to 10)
        headers = [cell.value for cell in input_sheet[6][1:10]]
        last_row = self._get_last_row(input_sheet, 'B')
        new_row = last_row + 1

        # Prepare data in the correct column order
        row_data = [data_dict.get(h, '') for h in headers]

        # Write the new row starting from column B (column index 2)
        for col_idx, value in enumerate(row_data, 2):
            input_sheet.cell(row=new_row, column=col_idx, value=value)

        wb.save(self.file_path)
        print(f"Successfully added new row at row {new_row}")
        return new_row

    def read_sales_forecast(self, as_dataframe=True):
        """
        Read forecast outputs from the 'Sales forecast' tab using openpyxl.
        """
        wb = openpyxl.load_workbook(self.file_path, data_only=True)
        forecast_sheet = wb['Sales forecast']

        last_row = self._get_last_row(forecast_sheet, 'B')

        # Data is in columns B, C, D (2, 3, 4)
        data = list(forecast_sheet.iter_rows(min_row=6, max_row=last_row, min_col=2, max_col=4, values_only=True))

        if as_dataframe:
            if len(data) < 2: # Need at least headers and one row of data
                return pd.DataFrame()
            headers = data[0]
            df = pd.DataFrame(data[1:], columns=headers)
            return df
        return data

# Example usage
if __name__ == "__main__":
    file_path = "microsoft_Sales forecast tracker small business.xlsx"
    backup_path = f"{file_path}.bak"

    # Create a backup if it doesn't exist, otherwise restore from it for a clean run
    if not os.path.exists(backup_path):
        shutil.copy(file_path, backup_path)
    else:
        shutil.copy(backup_path, file_path)

    model = SalesForecastModel(file_path)

    print("=" * 80)
    print("READING SALES FORECAST DATA (openpyxl)")
    print("=" * 80)
    forecast_df = model.read_sales_forecast()
    print(forecast_df.head())

    if not forecast_df.empty:
        # Use string casting to safely search for column names
        monthly_forecast_col = [col for col in forecast_df.columns if 'Monthly' in str(col)][0]
        cumulative_col = [col for col in forecast_df.columns if 'Cumulative' in str(col)][0]

        print(f"\nTotal Monthly Forecast: ${forecast_df[monthly_forecast_col].sum():,.2f}")
        print(f"Final Cumulative: ${forecast_df[cumulative_col].iloc[-1]:,.2f}")

    print("\n" + "=" * 80)
    print("ADDING NEW FORECAST INPUT ROW")
    print("=" * 80)
    new_data = {
        'Opportunity name': 'Agent Test Corp',
        'Sales \nagent': 'Jules',
        'Sales \nregion': 'US - West',
        'Sales \ncategory': 'Products',
        'Forecast amount': 50000,
        'Sales \nphase': 'Proposal',
        'Probability of sale': 0.75,
        'Forecast \nclose': datetime(2027, 1, 1),
        'Weighted forecast': 37500
    }
    new_row_num = model.add_forecast_input_row(new_data)
    print(f"Data added to row {new_row_num}. The spreadsheet has been updated.")
