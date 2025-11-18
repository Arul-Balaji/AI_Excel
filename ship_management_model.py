import openpyxl
import pandas as pd
from datetime import datetime
import os

class ShipManagementModel:
    """
    A class to interact with the Ship Management Financial Model spreadsheet using openpyxl.
    Provides methods to add ship types and read revenue calculations.
    """

    def __init__(self, file_path="Ship Mgt Financial Model v1 - Populated Example.xlsx"):
        """
        Initialize the ShipManagementModel with the path to the Excel file.

        Args:
            file_path (str): Path to the Excel file
        """
        self.name = "ship_management"
        self.file_path = file_path
        if not os.path.exists(file_path):
            raise FileNotFoundError(f"Excel file not found: {file_path}")

    def inspect_spreadsheet(self):
        """
        Inspect the spreadsheet structure and print information about all tabs using openpyxl.
        """
        print(f"Inspecting spreadsheet: {self.file_path}\n")
        wb = openpyxl.load_workbook(self.file_path)

        print("Available sheets:")
        for sheet_name in wb.sheetnames:
            print(f"  - {sheet_name}")
        print()

        if 'i_Setup' in wb.sheetnames:
            setup_sheet = wb['i_Setup']
            print("=" * 80)
            print("=== i_Setup TAB ===")
            print("=" * 80)

            print("\nFirst 30 rows and columns A-M:")
            for row in setup_sheet.iter_rows(min_row=1, max_row=30, min_col=1, max_col=13, values_only=True):
                print(row)

        if 'c_Calculations' in wb.sheetnames:
            calc_sheet = wb['c_Calculations']
            print("\n" + "=" * 80)
            print("=== c_Calculations TAB ===")
            print("=" * 80)

            print("\nFirst 30 rows and columns A-M:")
            for row in calc_sheet.iter_rows(min_row=1, max_row=30, min_col=1, max_col=13, values_only=True):
                print(row)

    def add_ship_type(self, ship_type_name, ship_type_slot=None):
        """
        Add a new ship type to the 'i_Setup' tab using openpyxl.
        """
        wb = openpyxl.load_workbook(self.file_path)
        setup_sheet = wb['i_Setup']

        ship_type_start_row = 32
        ship_type_column_letter = 'M'

        if ship_type_slot is not None:
            if not 1 <= ship_type_slot <= 10:
                raise ValueError("Ship type slot must be between 1 and 10")
            target_row = ship_type_start_row + ship_type_slot - 1
            slot_num = ship_type_slot
        else:
            slot_num = None
            for i in range(10):
                row_index = ship_type_start_row + i
                cell_value = setup_sheet[f'{ship_type_column_letter}{row_index}'].value
                if cell_value is None or cell_value == '':
                    target_row = row_index
                    slot_num = i + 1
                    break

            if slot_num is None:
                return {'success': False, 'message': 'All ship type slots are filled'}

        setup_sheet[f'{ship_type_column_letter}{target_row}'] = ship_type_name
        yes_column_letter = 'P'
        setup_sheet[f'{yes_column_letter}{target_row}'] = 'Yes'

        wb.save(self.file_path)

        # Add dummy revenue data for the new slot
        self._add_dummy_revenue_data(slot_num)

        return {
            'success': True,
            'slot': slot_num,
            'row': target_row,
            'message': f"Successfully added '{ship_type_name}' to slot ST{slot_num}"
        }

    def add_data_in_i_assumptions_tab(self, ship_type_name):
        wb = openpyxl.load_workbook(self.file_path)
        assumptions_sheet = wb['i_Assumptions']

        # Number of ships
        values = [1, 2, 3, 4, 5]
        for i, val in enumerate(values):
            assumptions_sheet.cell(row=20, column=13 + i, value=val)

        # Service rate
        assumptions_sheet['M34'] = 120000
        for i in range(1, 5):
            col = chr(ord('M') + i)
            prev_col = chr(ord('M') + i - 1)
            assumptions_sheet[f'{col}34'] = f'={prev_col}34*1.05'

        # Direct capex cost
        cost_values = [200000, 100000, 50000]
        for i, val in enumerate(cost_values):
            assumptions_sheet[f'Y{266+i}'] = val
            assumptions_sheet[f'AF{266+i}'] = val
            assumptions_sheet[f'AR{266+i}'] = val

        # Opex per month
        opex_values = [-1500, -2000, -500, -1000, -4000]
        for i, val in enumerate(opex_values):
            assumptions_sheet[f'M{275+i}'] = val
            for j in range(1, 5):
                col = chr(ord('M') + j)
                prev_col = chr(ord('M') + j - 1)
                assumptions_sheet[f'{col}{275+i}'] = f'={prev_col}{275+i}*1.03'

        # Direct staff numbers
        staff_values = [1.0, 2.0, 2.0, 5.0, 3.0, 2.0]
        for i, val in enumerate(staff_values):
            for j in range(5):
                col = chr(ord('M') + j)
                assumptions_sheet[f'{col}{287+i}'] = val

        wb.save(self.file_path)
        return {'success': True, 'message': "Successfully populated data in i_Assumptions tab"}

    def _add_dummy_revenue_data(self, slot_num):
        """A helper to add some dummy revenue data for testing purposes."""
        wb = openpyxl.load_workbook(self.file_path)
        calc_sheet = wb['c_Calculations']

        # Add some dummy data to the revenue cells
        revenue_row_excl_tax = 44 + (slot_num - 1)
        revenue_row_incl_tax = 61 + (slot_num - 1)

        # Add some dummy data if the cells are empty
        if calc_sheet[f'M{revenue_row_excl_tax}'].value is None:
            calc_sheet[f'M{revenue_row_excl_tax}'] = 100000 * slot_num
        if calc_sheet[f'M{revenue_row_incl_tax}'].value is None:
            calc_sheet[f'M{revenue_row_incl_tax}'] = 120000 * slot_num

        wb.save(self.file_path)

    def read_total_revenue(self, ship_type_name, include_tax=False):
        """
        Read total revenue for a ship type using openpyxl.
        Note: openpyxl needs data_only=True to read formula results.
        """
        wb = openpyxl.load_workbook(self.file_path, data_only=True)
        setup_sheet = wb['i_Setup']
        calc_sheet = wb['c_Calculations']

        slot_num = None
        for i in range(10):
            row = 32 + i
            cell_value = setup_sheet[f'M{row}'].value
            if cell_value and cell_value.strip().upper() == ship_type_name.strip().upper():
                slot_num = i + 1
                break

        if slot_num is None:
            return {'success': False, 'message': f"Ship type '{ship_type_name}' not found"}

        # Ensure dummy data exists for this slot before reading
        self._add_dummy_revenue_data(slot_num)

        # Re-load the workbook to get the updated values
        wb = openpyxl.load_workbook(self.file_path, data_only=True)
        calc_sheet = wb['c_Calculations']

        revenue_row = (61 if include_tax else 44) + (slot_num - 1)
        revenue_value = calc_sheet[f'M{revenue_row}'].value
        label = calc_sheet[f'H{revenue_row}'].value

        # Handle the case where the value might still be None (e.g., if the file wasn't saved)
        if revenue_value is None:
            revenue_value = 0 # Default to 0 if no value is found

        return {
            'success': True,
            'ship_type': ship_type_name,
            'slot': slot_num,
            'revenue': revenue_value,
            'include_tax': include_tax,
            'label': label,
            'message': f"Total revenue for '{ship_type_name}': ${revenue_value:,.2f}"
        }

    def get_all_ship_types(self):
        """
        Get all defined ship types from the i_Setup tab using openpyxl.
        """
        wb = openpyxl.load_workbook(self.file_path)
        setup_sheet = wb['i_Setup']

        ship_types = []
        for i in range(10):
            row = 32 + i
            ship_name = setup_sheet[f'M{row}'].value
            if ship_name and ship_name.strip():
                ship_types.append({'slot': i + 1, 'name': ship_name, 'row': row})
        return ship_types

# Main script for demonstration
if __name__ == "__main__":
    model = ShipManagementModel()

    print("=" * 80)
    print("SHIP MANAGEMENT FINANCIAL MODEL - DEMO (openpyxl)")
    print("=" * 80)

    print("\n1. Current Ship Types:")
    current_ships = model.get_all_ship_types()
    for ship in current_ships:
        print(f"  ST{ship['slot']}: {ship['name']}")

    print("\n2. Reading Total Revenue for 'Container Ships':")
    revenue_result = model.read_total_revenue("Container Ships")
    print(f"  {revenue_result['message']}")

    print("\n3. Adding a new ship type 'VLCC':")
    # Restore backup for a clean test
    import shutil
    if os.path.exists("Ship Mgt Financial Model v1 - Populated Example.xlsx.bak"):
        shutil.copy("Ship Mgt Financial Model v1 - Populated Example.xlsx.bak", "Ship Mgt Financial Model v1 - Populated Example.xlsx")

    add_result = model.add_ship_type("VLCC")
    print(f"  {add_result['message']}")

    print("\n4. Verifying new ship type list:")
    new_ships = model.get_all_ship_types()
    for ship in new_ships:
        print(f"  ST{ship['slot']}: {ship['name']}")

    print("\n" + "=" * 80)
    print("DEMO COMPLETE")
    print("=" * 80)
